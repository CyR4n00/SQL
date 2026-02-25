"""
db_engine.py - SQLite データベースエンジン
Excel / CSV のインポート・エクスポート・SQL実行を担当
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import sqlite3
import csv
import json
from pathlib import Path
from datetime import datetime


DB_PATH = Path(__file__).parent / "data" / "database.db"


class DBEngine:
    def __init__(self, db_path: Path = DB_PATH):
        db_path.parent.mkdir(exist_ok=True)
        self.db_path = db_path
        self.conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA journal_mode=WAL")
        self._init_meta()

    # ── 初期化 ────────────────────────────────────────────────
    def _init_meta(self):
        """インポート履歴テーブルを作成"""
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS _import_log (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                table_name TEXT,
                source    TEXT,
                rows      INTEGER,
                imported_at TEXT
            )
        """)
        self.conn.commit()

    # ── テーブル一覧 ──────────────────────────────────────────
    def list_tables(self) -> list[str]:
        cur = self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE '\\_%' ESCAPE '\\'")
        return [r[0] for r in cur.fetchall()]

    def table_info(self, table: str) -> list[dict]:
        cur = self.conn.execute(f"PRAGMA table_info('{table}')")
        return [dict(r) for r in cur.fetchall()]

    def row_count(self, table: str) -> int:
        cur = self.conn.execute(f"SELECT COUNT(*) FROM \"{table}\"")
        return cur.fetchone()[0]

    # ── CSV インポート ─────────────────────────────────────────
    def import_csv(self, filepath: str, table_name: str = None,
                   if_exists: str = "replace") -> dict:
        """
        CSV を SQLite テーブルにインポート
        if_exists: 'replace' | 'append' | 'fail'
        """
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"ファイルが見つかりません: {filepath}")

        table_name = table_name or _safe_name(path.stem)

        # CSV 読み込み（文字コード自動判定）
        rows, headers = _read_csv(path)

        if if_exists == "replace":
            self.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        elif if_exists == "fail":
            if table_name in self.list_tables():
                raise ValueError(f"テーブル '{table_name}' はすでに存在します")

        # テーブル作成（全カラム TEXT で受け入れ）
        cols_def = ", ".join(f'"{h}" TEXT' for h in headers)
        self.conn.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({cols_def})')

        # データ挿入
        placeholders = ", ".join("?" * len(headers))
        col_names    = ", ".join(f'"{h}"' for h in headers)
        self.conn.executemany(
            f'INSERT INTO "{table_name}" ({col_names}) VALUES ({placeholders})',
            [list(r.values()) for r in rows])

        # ログ記録
        self.conn.execute(
            "INSERT INTO _import_log (table_name, source, rows, imported_at) VALUES (?,?,?,?)",
            (table_name, str(path), len(rows), datetime.now().isoformat()))
        self.conn.commit()

        return {"table": table_name, "rows": len(rows), "columns": headers}

    # ── Excel インポート ──────────────────────────────────────
    def import_excel(self, filepath: str, sheet_name=None,
                     table_name: str = None, if_exists: str = "replace") -> list[dict]:
        """
        Excel ファイルをインポート。シートごとにテーブルを作成。
        sheet_name=None のとき全シートを処理
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl が必要です: pip install openpyxl")

        path = Path(filepath)
        wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)

        target_sheets = [sheet_name] if sheet_name else wb.sheetnames
        results = []

        for sname in target_sheets:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            tbl = table_name or _safe_name(f"{path.stem}_{sname}")

            rows_data = list(ws.values)
            if not rows_data:
                continue

            headers = [str(h) if h is not None else f"col_{i}"
                       for i, h in enumerate(rows_data[0])]
            data_rows = [
                [str(c) if c is not None else "" for c in row]
                for row in rows_data[1:]
            ]

            if if_exists == "replace":
                self.conn.execute(f'DROP TABLE IF EXISTS "{tbl}"')

            cols_def = ", ".join(f'"{h}" TEXT' for h in headers)
            self.conn.execute(f'CREATE TABLE IF NOT EXISTS "{tbl}" ({cols_def})')

            placeholders = ", ".join("?" * len(headers))
            col_names    = ", ".join(f'"{h}"' for h in headers)
            self.conn.executemany(
                f'INSERT INTO "{tbl}" ({col_names}) VALUES ({placeholders})',
                data_rows)

            self.conn.execute(
                "INSERT INTO _import_log (table_name, source, rows, imported_at) VALUES (?,?,?,?)",
                (tbl, f"{path}[{sname}]", len(data_rows), datetime.now().isoformat()))

            results.append({"table": tbl, "sheet": sname,
                             "rows": len(data_rows), "columns": headers})

        self.conn.commit()
        wb.close()
        return results

    # ── SQL 実行 ──────────────────────────────────────────────
    def execute_sql(self, sql: str) -> dict:
        """
        SQL を実行して結果を返す
        SELECT → {"type":"select", "columns":[], "rows":[], "count":N}
        その他 → {"type":"update", "affected":N}
        """
        sql_stripped = sql.strip().rstrip(";")
        try:
            cur = self.conn.execute(sql_stripped)
            if sql_stripped.upper().startswith("SELECT") or \
               sql_stripped.upper().startswith("WITH"):
                cols = [d[0] for d in cur.description] if cur.description else []
                rows = [list(r) for r in cur.fetchall()]
                return {"type": "select", "columns": cols,
                        "rows": rows, "count": len(rows)}
            else:
                self.conn.commit()
                return {"type": "update", "affected": cur.rowcount}
        except sqlite3.Error as e:
            raise RuntimeError(f"SQL エラー: {e}")

    # ── CSV エクスポート ──────────────────────────────────────
    def export_csv(self, sql: str, output_path: str,
                   encoding: str = "utf-8-sig") -> dict:
        """
        SQL の結果を CSV に書き出す
        utf-8-sig = Excel で文字化けしない BOM 付き UTF-8
        """
        result = self.execute_sql(sql)
        if result["type"] != "select":
            raise ValueError("SELECT 文のみエクスポートできます")

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, "w", newline="", encoding=encoding) as f:
            writer = csv.writer(f)
            writer.writerow(result["columns"])
            writer.writerows(result["rows"])

        return {"path": str(path), "rows": result["count"],
                "columns": result["columns"]}

    # ── Excel エクスポート ────────────────────────────────────
    def export_excel(self, queries: dict[str, str], output_path: str) -> dict:
        """
        { "シート名": "SQL文" } の辞書を受け取り、複数シートの Excel を出力
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl が必要です: pip install openpyxl")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # デフォルトシート削除

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(bold=True, color="FFFFFF", name="Arial")

        total_rows = 0
        for sheet_name, sql in queries.items():
            result = self.execute_sql(sql)
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel シート名 31文字制限

            # ヘッダー
            for ci, col in enumerate(result["columns"], 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font   = header_font
                cell.fill   = header_fill
                cell.alignment = Alignment(horizontal="center")

            # データ
            for ri, row in enumerate(result["rows"], 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val)

            # 列幅 自動調整
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            total_rows += result["count"]

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))

        return {"path": str(path), "sheets": list(queries.keys()),
                "total_rows": total_rows}

    # ── テーブル削除 ──────────────────────────────────────────
    def drop_table(self, table: str):
        self.conn.execute(f'DROP TABLE IF EXISTS "{table}"')
        self.conn.commit()

    # ── インポートログ ────────────────────────────────────────
    def import_log(self) -> list[dict]:
        cur = self.conn.execute(
            "SELECT * FROM _import_log ORDER BY id DESC LIMIT 50")
        return [dict(r) for r in cur.fetchall()]

    def close(self):
        self.conn.close()


# ── ヘルパー関数 ──────────────────────────────────────────────
def _safe_name(name: str) -> str:
    """テーブル名として安全な文字列に変換"""
    import re
    name = re.sub(r"[^\w]", "_", name)
    if name[0].isdigit():
        name = "t_" + name
    return name.lower()


def _read_csv(path: Path) -> tuple[list[dict], list[str]]:
    """文字コードを自動判定して CSV を読み込む"""
    for enc in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
        try:
            with open(path, encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                headers = list(rows[0].keys()) if rows else []
            return rows, headers
        except (UnicodeDecodeError, StopIteration):
            continue
    raise ValueError(f"文字コードを判定できませんでした: {path}")
